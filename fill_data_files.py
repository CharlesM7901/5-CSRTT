import openpyxl
from openpyxl import load_workbook
import pandas as pd

# Load file_names.xlsx as reference
file_names_df = pd.read_excel('file_names.xlsx')

# Extract date components from filename if not already done
if 'Year' not in file_names_df.columns:
    file_names_df[['Year', 'Month', 'Day', 'Hour', 'Minute', 'Second', 'Animal_Number']] = file_names_df['Name'].str.extract(
        r'(\d{4})_(\d{2})_(\d{2})__(\d{2})_(\d{2})_(\d{2})_(.+?)\.csv'
    )
    file_names_df[['Year', 'Month', 'Day']] = file_names_df[['Year', 'Month', 'Day']].astype(int)

# Define date mapping for each Jour
date_mapping = {
    'Jour 1': (2025, 12, 10),
    'Jour 2': (2025, 12, 11),
    'Jour 3': (2025, 12, 12),  # Already done, but including for completeness
    'Jour 4': (2025, 12, 14),
    'Jour 5': (2025, 12, 15),
    'Jour 6': (2025, 12, 16),
    'Jour 7': (2025, 12, 17),
    'Jour 8': (2025, 12, 18),
    'Jour 9': (2025, 12, 19),
    'Jour 10': (2025, 12, 20),
}

def get_animal_number(subject_id):
    """Extract just the digits from subject ID (remove V, *, etc.)"""
    return ''.join(c for c in str(subject_id) if c.isdigit())

def get_filename_for_subject_date(subject_id, year, month, day):
    """Find the filename matching the subject and date"""
    animal_number = get_animal_number(subject_id)
    
    match = file_names_df[(file_names_df['Year'] == year) & 
                          (file_names_df['Month'] == month) & 
                          (file_names_df['Day'] == day) & 
                          (file_names_df['Animal_Number'] == animal_number)]
    
    if len(match) > 0:
        return match.iloc[0]['Name']
    return None

# Load master_sheet_names.xlsx
wb = load_workbook('master_sheet_names.xlsx')

# Process each sheet
for sheet_name, (year, month, day) in date_mapping.items():
    try:
        ws = wb[sheet_name]
        
        # Process each row (skip header row 1)
        for row_idx in range(2, ws.max_row + 1):
            # Get the subject ID from the 'Sujets' column (column B)
            subject_id = ws.cell(row_idx, 2).value
            
            if subject_id:  # Only process if there's a subject ID
                filename = get_filename_for_subject_date(subject_id, year, month, day)
                # Set the Data_File value in column A
                ws.cell(row_idx, 1).value = filename
        
        print(f"✓ {sheet_name} ({year}/{month:02d}/{day:02d}): Filled {ws.max_row - 1} rows")
    except KeyError:
        print(f"✗ {sheet_name} not found in workbook")
    except Exception as e:
        print(f"✗ Error processing {sheet_name}: {str(e)}")

# Save the workbook
wb.save('master_sheet_names.xlsx')
print("\nAll Data_File columns filled and file saved successfully!")
